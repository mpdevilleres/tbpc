import openpyxl
from budget_mgt.models import Invoice
import pandas as pd

class InvoiceReport(object):

    template_path = r'/web/media/templates/invoice_cover.xlsx'
    output_row_pointer = 1
    reference = None
    date = None
    data = []
    column_dimensions = None
    row_dimensions = None
    total = {}
    def __init__(self, template_path=None):
        self.template = openpyxl.load_workbook(self.template_path if template_path is None else template_path)
        self.header_sheet = self.template.get_sheet_by_name('header')
        self.footer_sheet = self.template.get_sheet_by_name('footer')
        self.body_sheet = self.template.get_sheet_by_name('body')
        # make new workbook
        self.output_wb = openpyxl.Workbook()
        self.output_sheet = self.output_wb.create_sheet('report', index=0)

    def get_dimensions(self):
        self.column_dimensions = dict(**self.header_sheet.column_dimensions)

    def get_data(self):
        serial_no = 0
        # get data
        for record in Invoice.objects.filter(pk__in=[1,2,3,4,5]).all():
            serial_no += 1
            self.data.append(
                    {'B': serial_no,
                     'C': record.region,
                     'D': record.contractor.name,
                     'E': record.invoice_no,
                     'F': record.contract_no,
                     'G': record.revenue_amount,
                     'H': record.opex_amount,
                     'I': record.capex_amount,
                     'J': record.revenue_amount + record.opex_amount + record.capex_amount,
                     'K': record.task.task_no
                     }
                )
        df = pd.DataFrame.from_records(self.data)
        self.total['G'] = str(df['G'].sum())
        self.total['H'] = str(df['H'].sum())
        self.total['I'] = str(df['I'].sum())
        self.total['J'] = str(df['J'].sum())

        self.reference = 'Report-Reference-123313'
        self.date = 'Report Date'

    def make_header(self):
        # Add data to Header
        self.header_sheet['E2'] = self.reference
        self.header_sheet['E3'] = self.date

        # HEADER
        for row in self.header_sheet.iter_rows('A1:K6'):
            for cell in row:

                to_coor = '{}{}'.format(cell.column,self.output_row_pointer)
                to_cell = self.output_sheet.cell(to_coor)

                to_cell.value = cell.value
                to_cell.style = cell.style

            self.output_row_pointer += 1

        # Body
        # Copy table header
        for row in self.body_sheet.iter_rows('A2:K2'):
            for cell in row:

                to_coor = '{}{}'.format(cell.column,self.output_row_pointer)
                to_cell = self.output_sheet.cell(to_coor)

                to_cell.value = cell.value
                to_cell.style = cell.style

            self.output_row_pointer += 1

    def make_body(self):
        # populate table data
        # No.	Reg.	Contractor	Invoice No.	Contract	Revenue	OpEx	CapEx	Total Amt.	Budget/Yr.
        table_data_styles = {}
        for row in self.body_sheet.iter_rows('A3:K3'):
            for cell in row:
                table_data_styles[cell.column] = cell.style
        for record in self.data:
            for key in record.keys():
                to_coor = '{}{}'.format(key,self.output_row_pointer)
                to_cell = self.output_sheet.cell(to_coor)

                to_cell.value = record[key]
                to_cell.style = table_data_styles[key]

            self.output_row_pointer += 1

        # Copy table footer
        # Merge the Total Amount Cell
        merge_coor = '{1}{0}:{2}{0}'.format(self.output_row_pointer,'B', 'F')
        self.output_sheet.merge_cells(merge_coor)

        for row in self.body_sheet.iter_rows('A4:F4'):
            for cell in row:
                to_coor = '{}{}'.format(cell.column,self.output_row_pointer)
                to_cell = self.output_sheet.cell(to_coor)

                to_cell.value = cell.value
                to_cell.style = cell.style

            for column in ['G', 'H', 'I', 'J']:
                from_coor = '{}{}'.format(column,4)
                to_coor = '{}{}'.format(column,self.output_row_pointer)
                to_cell = self.output_sheet.cell(to_coor)

                to_cell.value = float(self.total[column])
                to_cell.style = self.body_sheet.cell(from_coor).style

            self.output_row_pointer += 1

    def make_footer(self):
        # Footer
        for row in self.footer_sheet.iter_rows('A2:K14'):
            for cell in row:

                to_coor = '{}{}'.format(cell.column,self.output_row_pointer)
                to_cell = self.output_sheet.cell(to_coor)

                to_cell.value = cell.value
                to_cell.style = cell.style

            self.output_row_pointer += 1

        for row in range(self.output_row_pointer,self.output_row_pointer-12,-1):
            self.output_sheet.row_dimensions[row].height = 30

        self.output_sheet.row_dimensions[self.output_row_pointer-13].height = 60

        merge_coor = '{1}{0}:{2}{0}'.format(self.output_row_pointer-13,'B', 'C')
        self.output_sheet.merge_cells(merge_coor)
        merge_coor = '{1}{0}:{2}{0}'.format(self.output_row_pointer-13,'D', 'J')
        self.output_sheet.merge_cells(merge_coor)

        for i in [1, 4, 5, 7, 10, 11]:
            merge_coor = '{1}{0}:{2}{0}'.format(self.output_row_pointer-i,'D', 'E')
            self.output_sheet.merge_cells(merge_coor)

        for i in [7, 10, 11]:
            merge_coor = '{1}{0}:{2}{0}'.format(self.output_row_pointer-i,'G', 'I')
            self.output_sheet.merge_cells(merge_coor)

    def run(self):

        self.get_data()
        self.make_header()
        self.make_body()
        self.make_footer()

        # Copying Dimensions Column
        for key, dimension in self.header_sheet.column_dimensions.items():
            self.output_sheet.column_dimensions[key].width = dimension.width

        self.output_wb.save('/web/media/templates/test1.xlsx')
